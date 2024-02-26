"""
Flight Test
William Raley
29 Jan 24

Test for location of flying crazyflie.
"""

# Python imports
import logging
import sys
import time
from threading import Event
from openpyxl import load_workbook

# Crazyflie imports
import cflib.crtp
from cflib.crazyflie import Crazyflie
from cflib.crazyflie.log import LogConfig
from cflib.crazyflie.syncCrazyflie import SyncCrazyflie
from cflib.positioning.motion_commander import MotionCommander
from cflib.positioning.position_hl_commander import PositionHlCommander
from cflib.utils import uri_helper

# Created Imports
import plotData

URI = uri_helper.uri_from_env(default='radio://0/80/2M/E7E7E7E7E7')
deck_atttached_event = Event()

logging.basicConfig(level=logging.ERROR)

DEFAULT_HEIGHT = 0.1
BOX_LIMIT = 2.6

position_estimate = [0 , 0 , 0]
#logconf = None

def createLog():
        # prepares log data
        logconf = LogConfig(name = 'Position', period_in_ms=10)
        logconf.add_variable('kalman.stateX' , 'float')
        logconf.add_variable('kalman.stateY' , 'float')
        logconf.add_variable('kalman.stateZ' , 'float')
        scf.cf.log.add_config(logconf)
        # prints position data
        logconf.data_received_cb.add_callback(log_pos_callback)
        # logs position data for graphing
        #logconf.data_received_cb.add_callback(log_stab_callback)

        return logconf

def setPos(logConf):
     print("Setting Position")
     logConf.start()
     time.sleep(0.1)
     logConf.stop()
     logConf.delete()
     return createLog()

def getPosTest(scf):
    with PositionHlCommander(scf, default_height = DEFAULT_HEIGHT) as pc:
            print(pc.get_position())
            
def holdTest(scf):

    print("Preparing To Fly")
    with PositionHlCommander(scf, default_height = DEFAULT_HEIGHT) as pc:
       
        pc._is_flying = False
        pc.take_off()
        pc.go_to(2.5 , 1.2, 1)
         # starts logging data
        logconf.start()
        trueLocation.append(2.5)
        trueLocation.append(1.2)
        trueLocation.append(0.)
        time.sleep(10)
        # ends data log
        logconf.stop()
        # landing sequence
        pc.down(.5)
        time.sleep(2)
        pc.land()      

def flyTest(scf):
     with PositionHlCommander(scf, default_height = DEFAULT_HEIGHT) as pc:

        # Sets Position
        pc._x = position_estimate[0]
        pc._y = position_estimate[1]
        pc._z = position_estimate[2]
        # Starts Log
        logConf.start()
        # Sets flying status to off
        pc._is_flying = False
       
        
        # Take Off
        print("Taking Off")
        pc.take_off(height=1, velocity=10)
        pc.go_to(2.4 , 1 , 1)
        time.sleep(1)

        # records Take off data
        recordData(fileName , "Take Off")

        time.sleep(5)

        # move
        print("Moving")
        # tracks what index timer starts at
        yLength.append(len(yLocation))
        # start timer
        start = time.time()
        pc.go_to(2.4 , 2.2 , 1)
        # stop timer
        stop = time.time()
        
        # Hold
        time.sleep(3)
    
        print("Landing")
        # landing sequence
        pc.down(.6)
        time.sleep(3)
        pc.land()

        return stop - start
   
def log_pos_callback(timestamp, data, logconf):
    #print(data)
    global position_estimate
    position_estimate[0] = data['kalman.stateX']
    position_estimate[1] = data['kalman.stateY']
    position_estimate[2] = data['kalman.stateZ']

def param_deck_flow(_, value_str):
    value = int(value_str)
    print(value)
    if value:
        deck_atttached_event.set()
        print('Deck is attached!')
    else:
         print('Deck is NOT attached!')

# Takes the information reported from data and puts it in the correct coord array
def log_stab_callback(timestamp, data, logconf):
   
   # converts data to string
   dataString = '%s\n'% data
   # spaces and periods are only allowed characters
   allowedString = '. '
   # removes all characters that are not numbers or in allowedString
   dataStringCut = ''.join(char for char in dataString if char.isdigit() or char in allowedString)
   # removes the excess .
   dataStringFinal = dataStringCut.replace(". ", "")

   # tracks what char in the data is currently on
   charCount = 0
   # tracks which coord is currently on
   coordCount = 0
 
   for char in dataStringFinal:
       # if decimal is found
       if char == ".":
           # create new string that contains the coord location
           coordString = dataStringFinal[charCount - 1: charCount + 5]
         
           if coordString == '0.0 0.':
               coordFloat = 0.0
           else:
                # creates float from string
                coordFloat = float(coordString)
           
           if coordCount == 0: # adds x coord
               xLocation.append(coordFloat)
           elif coordCount == 1: # adds y coord
               yLocation.append(coordFloat)
           else: # adds z coord
               zLocation.append(coordFloat)
           coordCount += 1 # increments to next coord
        # increments to next char
       charCount += 1
 
def recordData(fileName , sheet = None):
    
    doc = load_workbook(fileName)
    if sheet == None:
     sheet = doc.active
    else:
        sheet = doc.create_sheet(sheet)

    for i in range(len(xLocation)):
        data = [xLocation[i] , yLocation[i] , zLocation[i]]
        sheet.append(data)

    doc.save(fileName)

# Storage for location data
xLocation = []
yLocation = []
zLocation = []

yPath = []
yLength = []

if __name__ == '__main__':

    cflib.crtp.init_drivers()

    with SyncCrazyflie(URI, cf=Crazyflie(rw_cache='./cache')) as scf:
        print("Starting up")
        # checks to see if deck is attached 
        scf.cf.param.add_update_callback(group='deck', name='bcLoco', cb=param_deck_flow)

        time.sleep(1)

        logConf = createLog()        
    
        # if deck is not attached, run will fail 
        if not deck_atttached_event.wait(timeout=5):
            print('No Flow deck detected!')
            sys.exit(1)
        
        # sets controller and estimator
        scf.cf.param.set_value('stabilizer.controller' , '0')
        scf.cf.param.set_value('stabilizer.estimator' , '2')
        # sets file and sheet to have data sent to
        fileName = 'EKal_Auto.xlsx'
        
        #sheet = None
        
        logConf = setPos(logConf)
        # starts logging data
        logConf.data_received_cb.add_callback(log_stab_callback)
       
        #holdTest(scf)
        travelTime = flyTest(scf)
        #getPosTest(scf)
        
        # Stop Logging Data
        logConf.stop()
       
    
