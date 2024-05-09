"""
Flight Test
William Raley
29 Jan 24

Test for location of flying crazyflie.
"""

# Python imports
import logging
import sys
import time
from threading import Event
from openpyxl import Workbook
from openpyxl import load_workbook

# Crazyflie imports
import cflib.crtp
from cflib.crazyflie import Crazyflie
from cflib.crazyflie.log import LogConfig
from cflib.crazyflie.syncCrazyflie import SyncCrazyflie
from cflib.positioning.motion_commander import MotionCommander
from cflib.positioning.position_hl_commander import PositionHlCommander
from cflib.utils import uri_helper
from cflib.crazyflie.syncLogger import SyncLogger


URI = uri_helper.uri_from_env(default='radio://0/80/2M/E7E7E7E702')
deck_atttached_event = Event()

logging.basicConfig(level=logging.ERROR)

# Storage for location data

timeStamp = []
xLocation = []
yLocation = []
zLocation = []

tInitial = 0

position_estimate = [0 , 0 , 0]


# Creates log
def createLog():
        # prepares log data
        logconf = LogConfig(name = 'Position', period_in_ms=10)
       
        logconf.add_variable('kalman.stateX' , 'float')
        logconf.add_variable('kalman.stateY' , 'float')
        logconf.add_variable('kalman.stateZ' , 'float')
        """
        logconf.add_variable('stateEstimate.x' , 'float')
        logconf.add_variable('stateEstimate.y' , 'float')
        logconf.add_variable('stateEstimate.z' , 'float')
        """
        scf.cf.log.add_config(logconf)
        # prints position data
        logconf.data_received_cb.add_callback(log_pos_callback)
        #logconf.data_received_cb.add_callback(log_stab_callback)
        # logs position data for graphing
        #logconf.data_received_cb.add_callback(log_stab_callback)

        return logconf

# Sets the current position
def setPos(logConf):
     print("Setting Position")
     # Starts the log
     logConf.start()
     time.sleep(0.1)
     logConf.stop()
     logConf.delete()
    
     return createLog()

# Resets the location lists
def clearLocation(x, y, z):
        x.clear()
        y.clear()
        z.clear()
            
# Automates the flight of the CF
def flyTest(pc , xLocation , yLocation , zLocation , logConf , speed):

    # Sets Position
    pc._x = position_estimate[0]
    pc._y = position_estimate[1]
    pc._z = position_estimate[2]
    # height to land at
    lh = pc._y

    # Sets flying status to off
    pc._is_flying = False

    # Take Off
    print("Taking Off")
    #pc.take_off(height=1.0 , velocity = speed)
    pc.go_to(2.34 , 1.45 , 1.0 , speed)

    # records Take off data
    logConf.stop()
    recordData(fileName , xLocation , yLocation , zLocation)
    
    # resets point and log
    clearLocation(xLocation , yLocation , zLocation)
    logConf = createLog()  
    logConf.start()

    # move
    print("Moving")
    pc.go_to(2.3, 2.69 , 1.0 , speed)
    
    # Hold
    logConf.stop()
    recordData(fileName , xLocation , yLocation , zLocation ,"Move")

    # resets point and log
    clearLocation(xLocation , yLocation , zLocation)
    logConf = createLog()  
    logConf.start()

    
    # landing sequence
    print("Landing")
    pc.go_to(2.3, 2.69 , 0.3 , speed)
    pc._is_flying = True
    pc.land(velocity = speed)
    
def hold(pc , xLocation , yLocation , zLocation , logConf , speed):
     # Sets Position
    pc._x = position_estimate[0]
    pc._y = position_estimate[1]
    pc._z = position_estimate[2]
    

    # Sets flying status to off
    pc._is_flying = False

    # Take Off
    print("Taking Off")
    #pc.take_off(height=1.0 , velocity = speed)
    pc.up(.5)
    time.sleep(3)
    pc._is_flying = True
    pc.land(velocity = 0.1)

# Sets location to current estimate
def log_pos_callback(timestamp, data, logconf):
    
   
   
    position_estimate[0] = data['kalman.stateX']
    position_estimate[1] = data['kalman.stateY']
    position_estimate[2] = data['kalman.stateZ']
    """
    position_estimate[0] = data['stateEstimate.x']
    position_estimate[1] = data['stateEstimate.y']
    position_estimate[2] = data['stateEstimate.z']
    """

    timeStamp.append(timestamp)
    xLocation.append(position_estimate[0])
    yLocation.append(position_estimate[1])
    zLocation.append(position_estimate[2])   


# Checks to ensure deck is attached
def param_deck_flow(_, value_str):
    value = int(value_str)
    print(value)
    if value:
        deck_atttached_event.set()
        print('Deck is attached!')
    else:
         print('Deck is NOT attached!')

# Records data to spreadsheet
def recordData(fileName , x, y, z, sheet = None):
    
    doc = load_workbook(fileName)
    if sheet == None:
     sheet = doc.active
    else:
        sheet = doc.create_sheet(sheet)

    for i in range(len(x)):
        
        data = [x[i] , y[i] , z[i]]
        sheet.append(data)

    doc.save(fileName)

# Creates spreadsheet to record data
def createWB(fileName):
    wb = Workbook()
    wb.create_sheet("Take Off")
    del wb['Sheet']
    wb.save(fileName)

# Driver block of code
if __name__ == '__main__':

    cflib.crtp.init_drivers()

    with SyncCrazyflie(URI, cf=Crazyflie(rw_cache='./cache')) as scf:
        
    # print("Starting up")
        # checks to see if deck is attached 
        scf.cf.param.add_update_callback(group='deck', name='bcLoco', cb=param_deck_flow)

        time.sleep(1)

        firstPass = True

        # creates log for determining position
        logConf = createLog()        
    
        # if deck is not attached, run will fail 
        if not deck_atttached_event.wait(timeout=5):
            print('No Flow deck detected!')
            sys.exit(1)
        
        # sets controller and estimator
        scf.cf.param.set_value('stabilizer.controller' , '0')
        scf.cf.param.set_value('stabilizer.estimator' , '3')

        # Creates file for test
        fileName = '5.xlsx'
        createWB(fileName)
        
        # Sets speed
        speed = 0.5
        logConf = setPos(logConf)
        print(position_estimate)
        logConf.start()
        
        with PositionHlCommander(scf) as pc:
            #flyTest(pc , xLocation , yLocation , zLocation , logConf ,speed)
            hold(pc , xLocation , yLocation , zLocation , logConf , speed)
        # Stop Logging Data
        
        i = 0

        # Move by hand Test
        """
        while(1):
            if input() == "":
                logConf.stop()
               
                if i == 0:
                    recordData(fileName , xLocation , yLocation , zLocation)
                    i += 1
                elif i == 1:
                    recordData(fileName , xLocation , yLocation , zLocation ,"Move")
                    i += 1
                elif i == 2:
                    break
                
                # resets point and log
                clearLocation(xLocation , yLocation , zLocation)
                logConf = createLog()  
                logConf.start()
        """     
               


        logConf.stop()

       
    recordData(fileName , xLocation , yLocation , zLocation, "Land")
        
        
       
    
