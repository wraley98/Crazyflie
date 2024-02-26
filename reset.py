"""
Flight Test
William Raley
29 Jan 24

Test for location of flying crazyflie.
"""

# Python imports
import logging
import sys
import time
from threading import Event
from openpyxl import Workbook
from openpyxl import load_workbook

# Crazyflie imports
import cflib.crtp
from cflib.crazyflie import Crazyflie
from cflib.crazyflie.log import LogConfig
from cflib.crazyflie.syncCrazyflie import SyncCrazyflie
from cflib.positioning.motion_commander import MotionCommander
from cflib.positioning.position_hl_commander import PositionHlCommander
from cflib.utils import uri_helper


URI = uri_helper.uri_from_env(default='radio://0/80/2M/E7E7E7E7E7')
deck_atttached_event = Event()

logging.basicConfig(level=logging.ERROR)

DEFAULT_HEIGHT = 1

# Storage for location data

xLocation = []
yLocation = []
zLocation = []

position_estimate = [0 , 0 , 0]

#logconf = None

# Creates log
def createLog():
        # prepares log data
        logconf = LogConfig(name = 'Position', period_in_ms=10)
        
        logconf.add_variable('stateEstimate.x' , 'float')
        logconf.add_variable('stateEstimate.y' , 'float')
        logconf.add_variable('stateEstimate.z' , 'float')
        """
        logconf.add_variable('kalman.stateX' , 'float')
        logconf.add_variable('kalman.stateY' , 'float')
        logconf.add_variable('kalman.stateZ' , 'float')
        """
        scf.cf.log.add_config(logconf)
        # prints position data
        logconf.data_received_cb.add_callback(log_pos_callback)
        #logconf.data_received_cb.add_callback(log_stab_callback)
        # logs position data for graphing
        #logconf.data_received_cb.add_callback(log_stab_callback)

        return logconf

# Sets the current position
def setPos(logConf):
     print("Setting Position")
     # Starts the log
     logConf.start()
     time.sleep(0.1)
     logConf.stop()
     logConf.delete()

     return createLog()

# Resets the location lists
def clearLocation(x, y, z):
        x.clear()
        y.clear()
        z.clear()
            
# Automates the flight of the CF
def reset(pc , xLocation , yLocation , zLocation , logConf ):

    # Sets Position
    pc._x = position_estimate[0]
    pc._y = position_estimate[1]
    pc._z = position_estimate[2]
    
    # Sets flying status to off
    pc._is_flying = False
    
    # Take Off
    print("Taking Off")
    pc.take_off(height=.8, velocity = 5)
    pc.go_to(2.4 , 1 , .8 )
    pc.go_to(2.4, 1 , .75)
    time.sleep(2)
    pc.land(velocity = .1)
    

# Sets location to current estimate
def log_pos_callback(timestamp, data, logconf):
    #print(data)
    """
    position_estimate[0] = data['kalman.stateX']
    position_estimate[1] = data['kalman.stateY']
    position_estimate[2] = data['kalman.stateZ']
    """
    position_estimate[0] = data['stateEstimate.x']
    position_estimate[1] = data['stateEstimate.y']
    position_estimate[2] = data['stateEstimate.z']

    xLocation.append(position_estimate[0])
    yLocation.append(position_estimate[1])
    zLocation.append(position_estimate[2])   

# Checks to ensure deck is attached
def param_deck_flow(_, value_str):
    value = int(value_str)
    print(value)
    if value:
        deck_atttached_event.set()
        print('Deck is attached!')
    else:
         print('Deck is NOT attached!')

# Records data to spreadsheet
def recordData(fileName , x, y, z, sheet = None):
    
    doc = load_workbook(fileName)
    if sheet == None:
     sheet = doc.active
    else:
        sheet = doc.create_sheet(sheet)

    for i in range(len(x)):
        data = [x[i] , y[i] , z[i]]
        sheet.append(data)

    doc.save(fileName)

# Creates spreadsheet to record data
def createWB(fileName):
    wb = Workbook()
    wb.create_sheet("Take Off")
    del wb['Sheet']
    wb.save(fileName)

# Driver block of code
if __name__ == '__main__':

    cflib.crtp.init_drivers()

    with SyncCrazyflie(URI, cf=Crazyflie(rw_cache='./cache')) as scf:
        
    # print("Starting up")
        # checks to see if deck is attached 
        scf.cf.param.add_update_callback(group='deck', name='bcLoco', cb=param_deck_flow)


        # creates log for determining position
        logConf = createLog()        
    
        # if deck is not attached, run will fail 
        if not deck_atttached_event.wait(timeout=5):
            print('No Flow deck detected!')
            sys.exit(1)
        
        # sets controller and estimator
        scf.cf.param.set_value('stabilizer.controller' , '0')
        scf.cf.param.set_value('stabilizer.estimator' , '3')
        
        logConf.start()
        time.sleep(2)
        
        with PositionHlCommander(scf, default_height = DEFAULT_HEIGHT) as pc:
            reset(pc , xLocation , yLocation , zLocation , logConf)
       
        
        # Stop Logging Data
        logConf.stop()



        
        
       
    
